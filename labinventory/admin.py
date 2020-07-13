import os
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from labinventory.models import PressureGaugeUsageRecord, PressureGauge, GaugeType, Alarm, Person, UsageLocation


class PressureGaugeUsageRecordInline(admin.TabularInline):
    model = PressureGaugeUsageRecord


class PressureGaugeAdmin(admin.ModelAdmin):
    search_fields = ('number', 'gauge__company', 'gauge__mode', 'gauge__type', 'usage_location__name')
    list_display = ('__str__', 'gauge', 'usage_location')
    list_filter = ('gauge__mode', 'gauge__company', 'usage_location__name')
    save_on_top = True
    ordering = ['number']
    actions = ['excel_export']

    inlines = [PressureGaugeUsageRecordInline]

    def excel_export(self, request, queryset):
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pressure_gauges.xslx"'

        template_path = os.path.abspath(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pressure_gauges_template.xlsx'))

        wb = load_workbook(template_path)
        ws1 = wb.active

        cis = queryset
        row_offset = 2
        i = 0
        for ci in cis:
            _ = ws1.cell(column=1, row=i + row_offset, value='{}'.format(ci.number))
            _ = ws1.cell(column=2, row=i + row_offset, value='{}'.format(ci.gauge.company))
            _ = ws1.cell(column=3, row=i + row_offset, value='{}'.format(ci.gauge.get_modus_display()))
            _ = ws1.cell(column=4, row=i + row_offset, value='{}'.format(ci.usage_location))
            i += 1

        response.write(save_virtual_workbook(wb))
        return response


class AlarmAdmin(admin.ModelAdmin):
    list_display = ('type',)


admin.site.register(PressureGauge, PressureGaugeAdmin)
admin.site.register(GaugeType)
admin.site.register(Person)
admin.site.register(UsageLocation)
admin.site.register(PressureGaugeUsageRecord)
admin.site.register(Alarm, AlarmAdmin)
