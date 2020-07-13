from django.contrib import admin, messages
from cheminventory.models import Chemical, ChemicalInstance, GasCylinder, StorageLocation, GHS_H, \
    GHS_P, UsageLocation, Person, GasCylinderUsageRecord
from django.http import HttpResponseRedirect, HttpResponse
from django.template.loader import get_template
from django.template import Context
import os
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook


# Register your models here.

class InstanceStorageLocationListFilter(admin.SimpleListFilter):
    # Human-readable title which will be displayed in the
    # right admin sidebar just above the filter options.
    title = 'storage location'

    # Parameter for the filter that will be used in the URL query.
    parameter_name = 'storage_location'

    def lookups(self, request, model_admin):
        """
        We add the option to filter by 'all but waste' in order to get an
        overview of what is currently in stock in an easy way
        """
        lookups = [
            ('all_except_waste', u'All except waste'),
        ]

        # add all storage locations, because one should be able to filter by those as well
        locations = StorageLocation.objects.all()
        for location in locations:
            lookups.append((location.id, location.name))

        # sort the new list alphabetically, because easy.
        lookups = set(lookups)
        lookups = sorted(lookups, key=lambda tup: tup[1])
        # expects a tuple returned, so convert list to tuple
        return lookups

    def queryset(self, request, queryset):
        """
        Returns the filtered queryset based on the value
        provided in the query string and retrievable via
        `self.value()`.
        """
        # Compare the requested value (either '80s' or '90s')
        # to decide how to filter the queryset.
        if self.value() == None:
            return queryset
        elif self.value() == 'all_except_waste':
            waste_locations = ['Chemical Waste, liquid', 'Chemical Waste, solid']
            return queryset.exclude(storage_location__name__in=waste_locations)
        else:
            return queryset.filter(storage_location__exact=self.value())


class ChemicalInstanceAdmin(admin.ModelAdmin):
    def get_form(self, request, obj=None, **kwargs):
        form = super(ChemicalInstanceAdmin, self).get_form(request, obj, **kwargs)
        # form class is created per request by modelform_factory function
        # so it's safe to modify
        # we modify the the queryset to not show gases
        form.base_fields['chemical'].queryset = form.base_fields['chemical'].queryset.exclude(state_of_matter='GAS')
        return form

    search_fields = ('chemical__name', 'chemical__chemical_formula', 'chemical__cas', 'chemical__inchi')
    list_display = ('__str__', 'cas', 'state_of_matter', 'irritant', 'toxic', 'explosive', 'oxidizing', 'flammable',
                    'health_hazard', 'corrosive', 'environmentally_damaging')
    list_filter = (
        InstanceStorageLocationListFilter, 'chemical__state_of_matter', 'chemical__irritant', 'chemical__toxic',
        'chemical__explosive', 'chemical__oxidizing', 'chemical__flammable', 'chemical__health_hazard',
        'chemical__corrosive', 'chemical__environmentally_damaging', 'group')
    save_on_top = True
    # raw_id_fields = ('chemical', )
    ordering = ['chemical__name']
    actions = ['excel_export']

    def excel_export(self, request, queryset):
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="chemicals.xslx"'

        template_path = os.path.abspath(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chemistry_template.xlsx'))

        wb = load_workbook(template_path)
        ws1 = wb.active

        cis = queryset
        row_offset = 2
        i = 0
        for ci in cis:
            _ = ws1.cell(column=1, row=i + row_offset, value='{}'.format(ci.chemical.name))
            _ = ws1.cell(column=2, row=i + row_offset, value='{}'.format(ci.chemical.cas))
            _ = ws1.cell(column=4, row=i + row_offset, value='{}'.format(ci.company))
            _ = ws1.cell(column=5, row=i + row_offset, value='{}'.format(ci.item_number))
            amount, unit = '', ''
            for char in ci.quantity:
                if char.isalpha():
                    unit += char
                else:
                    amount += char
            _ = ws1.cell(column=6, row=i + row_offset, value='{}'.format(amount))
            _ = ws1.cell(column=7, row=i + row_offset, value='{}'.format(unit))

            _ = ws1.cell(column=8, row=i + row_offset, value='1')
            _ = ws1.cell(column=9, row=i + row_offset, value='{}'.format(ci.chemical.state_of_matter))
            _ = ws1.cell(column=10, row=i + row_offset, value='{}'.format(ci.storage_location))
            _ = ws1.cell(column=11, row=i + row_offset, value='{}'.format(ci.group))
            i += 1

        response.write(save_virtual_workbook(wb))
        return response


class GasCylinderUsageRecordInline(admin.TabularInline):
    model = GasCylinderUsageRecord


class GasCylinderAdmin(admin.ModelAdmin):
    search_fields = ('chemical__name', 'chemical__chemical_formula', 'chemical__cas', 'chemical__inchi')
    list_display = (
        '__str__', 'chemical__name', 'pressure', 'company', 'quantity', 'delivery_date', 'current_usage_location')
    list_filter = ('group', 'quantity', 'company', 'delivery_date')
    save_on_top = True
    raw_id_fields = ('chemical',)
    ordering = ['chemical__name']

    inlines = [GasCylinderUsageRecordInline]

    actions = ['QRcode', 'excel_export']

    def excel_export(self, request, queryset):
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="gas-cylinders.xslx"'

        template_path = os.path.abspath(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chemistry_template.xlsx'))

        wb = load_workbook(template_path)
        ws1 = wb.active

        cis = queryset
        row_offset = 2
        i = 0
        for ci in cis:
            _ = ws1.cell(column=1, row=i + row_offset, value='{}'.format(ci.chemical.name))
            _ = ws1.cell(column=2, row=i + row_offset, value='{}'.format(ci.chemical.cas))
            _ = ws1.cell(column=4, row=i + row_offset, value='{}'.format(ci.company))
            _ = ws1.cell(column=5, row=i + row_offset, value='')
            _ = ws1.cell(column=6, row=i + row_offset, value='{}'.format(ci.quantity))
            _ = ws1.cell(column=7, row=i + row_offset, value='L')

            _ = ws1.cell(column=8, row=i + row_offset, value='1')
            _ = ws1.cell(column=9, row=i + row_offset, value='Gas')
            _ = ws1.cell(column=10, row=i + row_offset, value='{}'.format(ci.storage_location))
            _ = ws1.cell(column=11, row=i + row_offset, value='{}'.format(ci.group))
            i += 1

        response.write(save_virtual_workbook(wb))
        return response

    def get_changeform_initial_data(self, request):
        new_id = GasCylinder.get_lowest_free_number()
        return {'cylinder_number': new_id}

    def QRcode(self, request, queryset):
        if len(queryset) == 1:
            s = queryset.get()
        else:
            return messages.error(request, 'Select only one cylinder!')
        return HttpResponseRedirect('/cheminventory/qrcode/%s' % (str(s.id)))


class ChemicalAdmin(admin.ModelAdmin):
    search_fields = ('cas', 'name', 'chemical_formula', 'inchi', 'csid', 'inchikey')
    list_display = (
        'name', 'chemical_formula', 'cas', 'csid', 'state_of_matter', 'irritant', 'toxic', 'explosive', 'oxidizing',
        'flammable', 'health_hazard', 'corrosive', 'environmentally_damaging')
    list_filter = (
        'state_of_matter', 'irritant', 'toxic', 'explosive', 'oxidizing', 'flammable', 'health_hazard', 'corrosive',
        'environmentally_damaging')
    save_on_top = True
    filter_horizontal = ('ghs_p', 'ghs_h')

    actions = ['new_instance']

    def new_instance(self, request, queryset):
        if len(queryset) == 1:
            s = queryset.get()
        else:
            return messages.error(request, 'Select only one chemical!')
        return HttpResponseRedirect('/admin/cheminventory/chemicalinstance/add/?chemical=%s' % (str(s.id)))


class ChemicalInstanceInline(admin.TabularInline):
    model = ChemicalInstance


class UsageLocationAdmin(admin.ModelAdmin):
    actions = ['print_doorsign']

    inlines = [ChemicalInstanceInline]

    def print_doorsign(self, request, queryset):
        if len(queryset) == 1:
            s = queryset.get()
        else:
            return messages.error(request, 'Select only one location!')
        return HttpResponseRedirect('/cheminventory/doorsign/%s' % (str(s.id)))


class StorageLocationAdmin(admin.ModelAdmin):
    actions = ['print_chemwaste']

    def print_chemwaste(self, request, queryset):
        if len(queryset) == 1:
            s = queryset.get()
        else:
            return messages.error(request, 'Select only one location!')
        return HttpResponseRedirect('/cheminventory/chemwaste/%s' % (str(s.id)))


class GHS_H_Admin(admin.ModelAdmin):
    save_on_top = True
    actions = ['show_all_chemicals']

    def show_all_chemicals(self, request, queryset):
        ghsids = queryset.values_list('id', flat=True)
        ghs = queryset.all()
        chemicals = ChemicalInstance.objects.filter(chemical__ghs_h__in=ghsids)
        t = get_template('cheminventory/ghs_list.html')
        c = Context({'ghs': ghs, 'chems': chemicals})

        html = t.render(c)
        return HttpResponse(html)


class GHS_P_Admin(admin.ModelAdmin):
    save_on_top = True
    actions = ['show_all_chemicals']

    def show_all_chemicals(self, request, queryset):
        ghsids = queryset.values_list('id', flat=True)
        ghs = queryset.all()
        chemicals = ChemicalInstance.objects.filter(chemical__ghs_p__in=ghsids)
        t = get_template('cheminventory/ghs_list.html')
        c = Context({'ghs': ghs, 'chems': chemicals})

        html = t.render(c)
        return HttpResponse(html)


class GasCylinderUsageRecordAdmin(admin.ModelAdmin):
    search_fields = (
    'usage_location__name', 'user__name', 'date', 'gas_cylinder__cylinder_number', 'gas_cylinder__chemical__name')
    list_display = ('gas_cylinder', 'usage_location', 'user', 'date')
    list_filter = ('usage_location', 'user', 'date', 'gas_cylinder')

    save_on_top = True


admin.site.register(Chemical, ChemicalAdmin)
admin.site.register(ChemicalInstance, ChemicalInstanceAdmin)
admin.site.register(GasCylinder, GasCylinderAdmin)
admin.site.register(StorageLocation, StorageLocationAdmin)
admin.site.register(GHS_H, GHS_H_Admin)
admin.site.register(GHS_P, GHS_P_Admin)
admin.site.register(UsageLocation, UsageLocationAdmin)
admin.site.register(Person)
admin.site.register(GasCylinderUsageRecord, GasCylinderUsageRecordAdmin)
